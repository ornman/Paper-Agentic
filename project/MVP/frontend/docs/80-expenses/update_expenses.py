"""更新项目开销记录"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

wb = load_workbook('D:/同步/project/frontend-prototype/docs/80-expenses/项目开销.xlsx')
ws = wb['明细账']

# 样式
input_font = Font(color='0000FF')  # 蓝色 = 手动输入
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 新增开销数据
expenses = [
    ['2026-03-23', 'API', '智谱 GLM', 'CodingPlanMax 包年', 1190, '支付宝', 'Max 套餐，立省¥300'],
    ['2026-03-23', 'API', 'Kimi', 'Allegro 包年', 6708, '支付宝', '¥559/月×12，立省¥1680'],
    ['2026-03-23', 'API', 'Minimax', 'CodingPlanMax 包年', 4502.4, '支付宝', 'Max 套餐，连续包年8折'],
]

# 找到最后一行数据（跳过合计行）
last_row = 2
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=1).value:
        last_row = row

# 插入新数据
for i, expense in enumerate(expenses):
    row_idx = last_row + i + 1
    for col_idx, value in enumerate(expense, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = input_font if col_idx == 5 else Font()
        cell.border = border
        if col_idx == 5:
            cell.alignment = Alignment(horizontal='right')

# 更新合计行
total_row = last_row + len(expenses) + 2
ws.cell(row=total_row, column=4, value='合计:').font = Font(bold=True)
ws.cell(row=total_row, column=5, value=f'=SUM(E2:E{total_row-1})')
ws.cell(row=total_row, column=5).font = Font(bold=True)
ws.cell(row=total_row, column=5).alignment = Alignment(horizontal='right')

# 更新项目信息 - 预算
ws_info = wb['项目信息']
ws_info['B5'] = 15000  # 预算设为 15000
ws_info['B5'].font = input_font

# 更新分类汇总 - 添加 API 类别
ws_cat = wb['分类汇总']
ws_cat['B3'] = '=SUMIF(明细账!B:B,"API",明细账!E:E)'

wb.save('D:/同步/project/frontend-prototype/docs/80-expenses/项目开销.xlsx')
print('已更新开销记录')

# 计算并显示总计
total = sum(e[4] for e in expenses)
print(f'本次添加: ¥{total}')
print(f'  - 智谱 GLM CodingPlanMax: ¥1,190')
print(f'  - Kimi Allegro: ¥6,708')
print(f'  - Minimax CodingPlanMax: ¥4,502.4')
