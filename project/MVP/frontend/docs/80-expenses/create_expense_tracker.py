"""创建项目开销记账单"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# === Sheet 1: 明细账 ===
ws = wb.active
ws.title = "明细账"

# 样式定义
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4472C4")
input_font = Font(color="0000FF")  # 蓝色 = 手动输入
formula_font = Font(color="000000")  # 黑色 = 公式
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')

# 表头
headers = ["日期", "类别", "项目", "说明", "金额(元)", "支付方式", "备注"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# 列宽
widths = [12, 10, 15, 25, 12, 12, 20]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# 示例数据
sample_data = [
    ["2026-03-20", "开发", "UI原型", "Vue 前端开发", 0, "自研", "纯前端展示"],
    ["2026-03-21", "工具", "WPS插件", "wpsjs 调试环境", 0, "自研", ""],
]
for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = input_font if col_idx == 5 else formula_font
        cell.border = border
        if col_idx == 5:  # 金额列右对齐
            cell.alignment = Alignment(horizontal='right')

# 合计行
total_row = len(sample_data) + 3
ws.cell(row=total_row, column=4, value="合计:").font = Font(bold=True)
ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row-1})")
ws.cell(row=total_row, column=5).font = formula_font
ws.cell(row=total_row, column=5).alignment = Alignment(horizontal='right')

# === Sheet 2: 分类汇总 ===
ws2 = wb.create_sheet("分类汇总")

categories = ["开发", "工具", "API", "设计", "服务器", "其他"]
ws2.cell(row=1, column=1, value="类别").font = header_font
ws2.cell(row=1, column=1).fill = header_fill
ws2.cell(row=1, column=2, value="金额(元)").font = header_font
ws2.cell(row=1, column=2).fill = header_fill
ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 12

for i, cat in enumerate(categories, 2):
    ws2.cell(row=i, column=1, value=cat)
    ws2.cell(row=i, column=2, value=f'=SUMIF(明细账!B:B,"{cat}",明细账!E:E)')

ws2.cell(row=len(categories)+3, column=1, value="总计:").font = Font(bold=True)
ws2.cell(row=len(categories)+3, column=2, value=f"=SUM(B2:B{len(categories)+1})")

# === Sheet 3: 项目信息 ===
ws3 = wb.create_sheet("项目信息")
info = [
    ["项目名称", "WPS 论文写作辅助工具 - UI 原型"],
    ["创建日期", datetime.now().strftime("%Y-%m-%d")],
    ["状态", "已冻结"],
    ["负责人", ""],
    ["预算(元)", 0],
    ["已花费(元)", "=分类汇总!B9"],
    ["剩余(元)", "=E5-E6"],
]
for row_idx, (label, value) in enumerate(info, 1):
    ws3.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
    ws3.cell(row=row_idx, column=2, value=value)
    if isinstance(value, str) and value.startswith("="):
        ws3.cell(row=row_idx, column=2).font = formula_font
    else:
        ws3.cell(row=row_idx, column=2).font = input_font

ws3.column_dimensions['A'].width = 15
ws3.column_dimensions['B'].width = 35

output_path = "D:/同步/project/frontend-prototype/docs/80-expenses/项目开销.xlsx"
wb.save(output_path)
print(f"已创建: {output_path}")
